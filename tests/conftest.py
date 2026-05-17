"""Shared pytest fixtures + reporting hooks for the BioPipelines test suite."""

import csv
import os
from pathlib import Path

import pytest


FIXTURES_DIR = Path(__file__).parent / "fixtures"
REPO_ROOT = Path(__file__).resolve().parent.parent
# Write results to tests/_results/ (gitignored). Previously written into
# tests/ directly, where the .csv/.xlsx ended up tracked and every run
# dirtied the working tree.
RESULTS_DIR = Path(__file__).resolve().parent / "_results"
RESULTS_CSV = RESULTS_DIR / "test_results.csv"
RESULTS_XLSX = RESULTS_DIR / "test_results.xlsx"

_CASE_STASH_KEY = pytest.StashKey[dict]()


# ── active-pipeline isolation ─────────────────────────────────────────────────

@pytest.fixture(autouse=True)
def _reset_active_pipeline():
    """Ensure the module-level ``_active_pipeline`` context is empty between
    tests. ``on_the_fly=True`` pipelines constructed in a test without a
    ``with`` block (notably shell-safety tests) leak the active context and
    make any subsequent test that instantiates a tool return a Standardized-
    Output or raise 'Cannot nest Pipeline contexts'."""
    try:
        from biopipelines.pipeline import _active_pipeline
    except ImportError:
        yield
        return
    _active_pipeline.set(None)
    yield
    _active_pipeline.set(None)


# ── local_config / isolated_cwd fixtures ──────────────────────────────────────

@pytest.fixture
def local_config(monkeypatch, tmp_path):
    """Point ConfigManager at tests/fixtures/config.local.yaml and reset its
    singleton state so each test gets a fresh load.

    Yields the absolute path of the fixture config file.
    """
    from biopipelines.config_manager import ConfigManager

    config_path = FIXTURES_DIR / "config.local.yaml"
    assert config_path.exists(), f"Missing fixture: {config_path}"

    ConfigManager._instance = None
    ConfigManager._config = None
    ConfigManager._variant = None

    monkeypatch.setattr(
        ConfigManager, "_get_config_path",
        classmethod(lambda cls, variant=None: str(config_path)),
    )

    yield config_path

    ConfigManager._instance = None
    ConfigManager._config = None
    ConfigManager._variant = None


@pytest.fixture
def slurm_local_config(monkeypatch, tmp_path):
    """Like ``local_config`` but pointing at ``config.slurm_local.yaml`` so the
    framework's SLURM script-generation path is exercised. Tests that need to
    inspect ``slurm_batch*.sh`` artifacts use this instead of ``local_config``."""
    from biopipelines.config_manager import ConfigManager

    config_path = FIXTURES_DIR / "config.slurm_local.yaml"
    assert config_path.exists(), f"Missing fixture: {config_path}"

    ConfigManager._instance = None
    ConfigManager._config = None
    ConfigManager._variant = None

    monkeypatch.setattr(
        ConfigManager, "_get_config_path",
        classmethod(lambda cls, variant=None: str(config_path)),
    )

    yield config_path

    ConfigManager._instance = None
    ConfigManager._config = None
    ConfigManager._variant = None


@pytest.fixture
def isolated_cwd(tmp_path, monkeypatch):
    """Run a test with cwd set to an isolated temp directory."""
    monkeypatch.chdir(tmp_path)
    return tmp_path


@pytest.fixture
def new_pipeline():
    """Factory for a minimal local-config Pipeline used by smoke tests."""
    from biopipelines.pipeline import Pipeline

    def _make(job: str):
        return Pipeline(
            project="TestSuite",
            job=job,
            description=f"Smoke test: {job}",
            on_the_fly=False,
            local_output=True,
            config="local",
        )
    return _make


@pytest.fixture
def new_slurm_pipeline():
    """Factory for a minimal SLURM-config Pipeline. Needs ``slurm_local_config``
    to be active so ConfigManager loads the SLURM-flavoured fixture."""
    from biopipelines.pipeline import Pipeline

    def _make(job: str):
        return Pipeline(
            project="TestSuite",
            job=job,
            description=f"Smoke test: {job}",
            on_the_fly=False,
            local_output=True,
            config="slurm_local",
        )
    return _make


@pytest.fixture
def assert_valid_script():
    """Assert that a saved pipeline.sh exists and contains expected markers."""
    import os as _os

    def _check(script_path: str, *markers: str):
        assert _os.path.isfile(script_path), f"pipeline.sh missing: {script_path}"
        content = open(script_path, encoding="utf-8").read()
        assert content.startswith("#!/bin/bash"), "missing shebang"
        assert len(content) > 200, "script suspiciously short"
        for marker in markers:
            assert marker in content, f"expected marker {marker!r} missing from script"
    return _check


# ── record_case fixture (adds input/expected/actual to the report) ────────────

@pytest.fixture
def record_case(request):
    """Let a test record its (input, expected, actual) so the XLSX report can
    show what was tested, not just the pass/fail flag.

    Usage inside a test:
        def test_foo(record_case):
            record_case(input="a_<0..1>", expected=["a_0", "a_1"], actual=actual)
            assert actual == ["a_0", "a_1"]

    For parametrized tests, call this once inside the test body.
    """
    def _record(*, input, expected, actual):
        node = request.node
        node.stash[_CASE_STASH_KEY] = {
            "input": _short_repr(input),
            "expected": _short_repr(expected),
            "actual": _short_repr(actual),
            "matched": expected == actual,
        }
    return _record


def _short_repr(value, limit: int = 200) -> str:
    r = repr(value)
    if len(r) > limit:
        return r[: limit - 3] + "..."
    return r


# ── per-test result capture + post-session CSV/XLSX report ────────────────────

_RESULT_ROWS: list[dict] = []


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    """Capture each test's call-phase outcome, duration, and recorded case."""
    outcome = yield
    report = outcome.get_result()

    # Only log one row per test; prefer the call phase, but fall back to a
    # non-passing setup phase so collection/setup errors still show up.
    if report.when != "call" and not (report.when == "setup" and report.outcome != "passed"):
        return

    nodeid = report.nodeid
    parts = nodeid.split("::")
    module = parts[0] if parts else nodeid
    if len(parts) == 3:
        classname, testname = parts[1], parts[2]
    elif len(parts) == 2:
        classname, testname = "", parts[1]
    else:
        classname, testname = "", nodeid

    case = item.stash.get(_CASE_STASH_KEY, None) if hasattr(item, "stash") else None

    _RESULT_ROWS.append({
        "module": module,
        "class": classname,
        "test": testname,
        "outcome": report.outcome,
        "duration_s": round(report.duration, 4),
        "input": case["input"] if case else "",
        "expected": case["expected"] if case else "",
        "actual": case["actual"] if case else "",
        "matched": ("yes" if case["matched"] else "no") if case else "",
        "error_detail": _short_repr(report.longreprtext, 500) if report.outcome == "failed" else "",
        "phase": report.when,
        "nodeid": nodeid,
    })


def pytest_sessionfinish(session, exitstatus):
    """Write collected test outcomes to outputs/test_results.{csv,xlsx}."""
    if not _RESULT_ROWS:
        return

    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    columns = [
        "module", "class", "test", "outcome", "duration_s",
        "input", "expected", "actual", "matched",
        "error_detail", "phase", "nodeid",
    ]

    with RESULTS_CSV.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        writer.writerows(_RESULT_ROWS)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.append(columns)
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    outcome_fills = {
        "passed": PatternFill("solid", fgColor="C6EFCE"),
        "failed": PatternFill("solid", fgColor="FFC7CE"),
        "skipped": PatternFill("solid", fgColor="FFEB9C"),
        "error": PatternFill("solid", fgColor="FFC7CE"),
    }
    outcome_col_idx = columns.index("outcome") + 1

    wrap = Alignment(wrap_text=True, vertical="top")
    wrap_cols = {"input", "expected", "actual", "error_detail"}

    for row in _RESULT_ROWS:
        ws.append([row[c] for c in columns])
        r_idx = ws.max_row
        fill = outcome_fills.get(row["outcome"])
        if fill is not None:
            ws.cell(row=r_idx, column=outcome_col_idx).fill = fill
        for i, col in enumerate(columns, start=1):
            if col in wrap_cols:
                ws.cell(row=r_idx, column=i).alignment = wrap

    for i, col in enumerate(columns, start=1):
        max_len = max([len(str(r[col])) for r in _RESULT_ROWS] + [len(col)])
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len + 2, 60)

    ws.freeze_panes = "A2"

    # Summary sheet with counts per outcome.
    summary = wb.create_sheet("summary")
    summary.append(["outcome", "count"])
    for cell in summary[1]:
        cell.font = header_font
    counts: dict[str, int] = {}
    for r in _RESULT_ROWS:
        counts[r["outcome"]] = counts.get(r["outcome"], 0) + 1
    for outcome, count in sorted(counts.items()):
        summary.append([outcome, count])
    summary.append([])
    summary.append(["recorded cases", sum(1 for r in _RESULT_ROWS if r["input"])])
    summary.append(["total tests", len(_RESULT_ROWS)])

    wb.save(RESULTS_XLSX)
