#!/usr/bin/env python3
"""
Runtime script for DNA encoding with codon optimization.

Reads protein sequences from CSV and generates DNA sequences optimized for
specified organism(s) using CoCoPUTs codon usage tables.

Usage: python pipe_dna_encoder.py --config <config.json>
"""

import argparse
import json
import pandas as pd
import random
from typing import Dict, List, Tuple
from itertools import combinations

# CoCoPUTs codon usage tables (Source: https://dnahive.fda.gov/dna.cgi?cmd=codon_usage)
COCOPUTS_TABLES = {
    "Homo sapiens": """TTT	17.14	(1385301)	TCT	16.93	(1368632)	TAT	12.11	( 978774)	TGT	10.40	( 841042)
TTC	17.48	(1413268)	TCC	17.32	(1399962)	TAC	13.49	(1090514)	TGC	10.81	( 873765)
TTA	 8.71	( 703680)	TCA	14.14	(1142684)	TAA	 0.44	(  35218)	TGA	 0.79	(  63801)
TTG	13.44	(1086777)	TCG	 4.03	( 325925)	TAG	 0.35	(  28499)	TGG	11.60	( 937286)

CTT	14.08	(1138433)	CCT	19.31	(1560898)	CAT	11.83	( 956479)	CGT	 4.55	( 367659)
CTC	17.81	(1439345)	CCC	19.11	(1544626)	CAC	14.65	(1184041)	CGC	 8.71	( 704401)
CTA	 7.44	( 601662)	CCA	18.92	(1529004)	CAA	14.06	(1136523)	CGA	 6.42	( 518818)
CTG	36.10	(2918400)	CCG	 6.22	( 503096)	CAG	35.53	(2872161)	CGG	10.79	( 871786)

ATT	16.48	(1331901)	ACT	14.26	(1152700)	AAT	18.43	(1489775)	AGT	14.05	(1135376)
ATC	18.67	(1508988)	ACC	17.85	(1442511)	AAC	18.30	(1478832)	AGC	19.69	(1591829)
ATA	 8.08	( 652939)	ACA	16.52	(1335468)	AAA	27.48	(2221062)	AGA	13.28	(1073213)
ATG	21.53	(1739992)	ACG	 5.59	( 452037)	AAG	31.77	(2567940)	AGG	12.13	( 980476)

GTT	11.74	( 949137)	GCT	18.99	(1534685)	GAT	24.03	(1942185)	GGT	10.83	( 875715)
GTC	13.44	(1086717)	GCC	25.84	(2088762)	GAC	24.27	(1961667)	GGC	19.79	(1599325)
GTA	 7.66	( 618960)	GCA	17.04	(1377145)	GAA	33.65	(2719693)	GGA	17.12	(1384137)
GTG	25.87	(2090923)	GCG	 5.91	( 477758)	GAG	39.67	(3206546)	GGG	15.35	(1240793)""",

    "Escherichia coli": """TTT	22.28	( 884859304)	TCT	 8.59	( 341230176)	TAT	16.28	( 646775431)	TGT	 5.15	( 204369628)
TTC	16.26	( 645822743)	TCC	 8.82	( 350462264)	TAC	12.18	( 483602182)	TGC	 6.31	( 250671692)
TTA	13.72	( 544875130)	TCA	 7.50	( 297755617)	TAA	 2.01	(  79698775)	TGA	 1.00	(  39709081)
TTG	13.33	( 529575852)	TCG	 8.79	( 349261974)	TAG	 0.23	(   9277733)	TGG	15.22	( 604449078)

CTT	11.39	( 452555895)	CCT	 7.19	( 285670480)	CAT	12.77	( 507363784)	CGT	20.72	( 822947687)
CTC	10.95	( 434982409)	CCC	 5.54	( 219881409)	CAC	 9.44	( 374998113)	CGC	21.55	( 855993095)
CTA	 3.88	( 154295599)	CCA	 8.42	( 334296050)	CAA	15.03	( 596935496)	CGA	 3.64	( 144705009)
CTG	52.32	(2078211248)	CCG	22.80	( 905697146)	CAG	29.26	(1162220193)	CGG	 5.68	( 225733345)

ATT	30.19	(1199278030)	ACT	 8.98	( 356533130)	AAT	18.08	( 717967207)	AGT	 9.02	( 358422205)
ATC	24.68	( 980290510)	ACC	23.00	( 913523340)	AAC	21.50	( 853806657)	AGC	15.92	( 632395429)
ATA	 4.72	( 187294284)	ACA	 7.53	( 299007362)	AAA	33.81	(1342811432)	AGA	 2.35	(  93340146)
ATG	27.65	(1098018887)	ACG	14.52	( 576519827)	AAG	10.67	( 423892201)	AGG	 1.40	(  55660373)

GTT	18.38	( 729993910)	GCT	15.54	( 617228810)	GAT	32.42	(1287642772)	GGT	24.54	( 974821805)
GTC	15.12	( 600554770)	GCC	25.62	(1017548717)	GAC	19.23	( 763756668)	GGC	28.83	(1145183663)
GTA	10.96	( 435276257)	GCA	20.59	( 817719553)	GAA	39.62	(1573751166)	GGA	 8.37	( 332414382)
GTG	25.97	(1031522759)	GCG	32.96	(1309274393)	GAG	18.29	( 726478043)	GGG	11.27	( 447491989)""",

    "Saccharomyces cerevisiae": """TTT	26.26	(76936)	TCT	23.34	(68371)	TAT	19.08	( 55892)	TGT	 7.86	(23021)
TTC	17.90	(52428)	TCC	14.05	(41167)	TAC	14.60	( 42768)	TGC	 4.79	(14043)
TTA	26.32	(77090)	TCA	19.05	(55799)	TAA	 0.97	(  2843)	TGA	 0.61	( 1792)
TTG	26.48	(77571)	TCG	 8.72	(25535)	TAG	 0.47	(  1391)	TGG	10.36	(30352)

CTT	12.34	(36134)	CCT	13.57	(39748)	CAT	13.91	( 40756)	CGT	 6.28	(18389)
CTC	 5.55	(16266)	CCC	 6.92	(20272)	CAC	 7.76	( 22722)	CGC	 2.65	( 7770)
CTA	13.53	(39630)	CCA	17.79	(52123)	CAA	27.06	( 79280)	CGA	 3.12	( 9148)
CTG	10.68	(31271)	CCG	 5.44	(15931)	CAG	12.42	( 36397)	CGG	 1.85	( 5407)

ATT	30.10	(88169)	ACT	20.22	(59227)	AAT	36.55	(107068)	AGT	14.59	(42745)
ATC	16.98	(49726)	ACC	12.47	(36530)	AAC	24.77	( 72563)	AGC	 9.98	(29225)
ATA	18.33	(53698)	ACA	18.18	(53255)	AAA	42.82	(125427)	AGA	21.03	(61608)
ATG	20.71	(60656)	ACG	 8.16	(23913)	AAG	30.46	( 89219)	AGG	 9.47	(27737)

GTT	21.45	(62830)	GCT	20.24	(59286)	GAT	38.00	(111311)	GGT	22.53	(66011)
GTC	11.22	(32866)	GCC	12.12	(35517)	GAC	20.35	( 59621)	GGC	 9.79	(28677)
GTA	12.07	(35367)	GCA	16.26	(47638)	GAA	45.70	(133873)	GGA	11.19	(32778)
GTG	10.74	(31449)	GCG	 6.19	(18118)	GAG	19.53	( 57212)	GGG	 6.07	(17778)"""
}

# Organism code mapping
ORGANISM_MAP = {
    "EC": "Escherichia coli",
    "SC": "Saccharomyces cerevisiae",
    "HS": "Homo sapiens"
}


def parse_cocoputs_table(table_text: str) -> Dict[str, float]:
    """Parse CoCoPUTs table text into codon frequency dictionary."""
    frequency_table = {}
    lines = table_text.split('\n')
    for line in lines:
        line = line.strip()
        if line == "":
            continue
        tab_split = line.split('\t')
        codon_frequency = [(tab_split[i], tab_split[i + 1]) for i in [0, 3, 6, 9]]
        for codon, frequency_s in codon_frequency:
            frequency = float(frequency_s.replace('\xa0', ' '))
            frequency_table[codon] = frequency
    return frequency_table


def create_mixed_table(freq_tables: List[Dict[str, float]]) -> Dict[str, float]:
    """Create mixed frequency table using minimum frequencies."""
    mixed_table = {}
    codons = list(freq_tables[0].keys())
    for codon in codons:
        mixed_table[codon] = min(table[codon] for table in freq_tables)
    return mixed_table


def create_aa_codon_frequency_table(frequency_table: Dict[str, float]) -> Dict[str, Tuple[List[str], List[float]]]:
    """Create amino acid to codon frequency mapping."""
    # Standard genetic code
    from Bio.Data import CodonTable
    std_table = CodonTable.standard_dna_table.forward_table

    aa_codon_freq = {}
    for codon, aa in std_table.items():
        if aa not in aa_codon_freq:
            aa_codon_freq[aa] = ([], [])
        aa_codon_freq[aa][0].append(codon)
        aa_codon_freq[aa][1].append(frequency_table[codon])

    return aa_codon_freq


def encode_sequence_thresholded_weighted(sequence: str, aa_codon_freq: Dict[str, Tuple[List[str], List[float]]]) -> str:
    """
    Encode protein sequence to DNA using thresholded weighted method.

    For each amino acid:
    - Sample only from codons with frequency >= 10‰
    - If no codons >= 10‰, use codons >= 5‰
    - If no codons >= 5‰, use most frequent codon
    """
    dna_codons = []
    for aa in sequence:
        if aa not in aa_codon_freq:
            raise ValueError(f"Invalid amino acid: {aa}")

        codons, frequencies = aa_codon_freq[aa]

        # Try threshold of 10‰
        threshold_indices = [i for i in range(len(frequencies)) if frequencies[i] >= 10.0]
        if len(threshold_indices) > 0:
            codons_threshold = [codons[i] for i in threshold_indices]
            frequencies_threshold = [frequencies[i] for i in threshold_indices]
            chosen_codon = random.choices(codons_threshold, weights=frequencies_threshold, k=1)[0]
        else:
            # Try threshold of 5‰
            threshold_indices = [i for i in range(len(frequencies)) if frequencies[i] >= 5.0]
            if len(threshold_indices) > 0:
                codons_threshold = [codons[i] for i in threshold_indices]
                frequencies_threshold = [frequencies[i] for i in threshold_indices]
                chosen_codon = random.choices(codons_threshold, weights=frequencies_threshold, k=1)[0]
            else:
                # Use most frequent codon
                max_frequency = max(frequencies)
                chosen_codon = codons[frequencies.index(max_frequency)]

        dna_codons.append(chosen_codon)

    return ''.join(dna_codons)


def main():
    parser = argparse.ArgumentParser(description='DNA encoding with codon optimization')
    parser.add_argument('--config', type=str, required=True, help='Path to config JSON file')
    args = parser.parse_args()

    # Load configuration
    with open(args.config, 'r') as f:
        config = json.load(f)

    sequences_csv = config['sequences_csv']
    organism_code = config['organism']
    dna_output = config['dna_output']
    excel_output = config['excel_output']
    info_output = config['info_output']

    print(f"DNA Encoder is based on CoCoPUTs (HIVE, tables updated April 2024). Please cite accordingly.")
    print(f"Reading sequences from: {sequences_csv}")
    print(f"Target organism(s): {organism_code}")

    # Load protein sequences
    sequences_df = pd.read_csv(sequences_csv)
    if 'sequence' not in sequences_df.columns or 'id' not in sequences_df.columns:
        raise ValueError("Input CSV must have 'id' and 'sequence' columns")

    # Parse organism code (e.g., "EC", "EC&HS", "EC&HS&SC")
    organism_codes = [code.strip() for code in organism_code.split('&')]
    organism_names = [ORGANISM_MAP[code] for code in organism_codes]

    print(f"Organisms: {', '.join(organism_names)}")

    # Parse frequency tables
    print("Parsing codon frequency tables...")
    frequency_tables = {}
    for org_name in organism_names:
        frequency_tables[org_name] = parse_cocoputs_table(COCOPUTS_TABLES[org_name])

    # Create mixed table if multiple organisms
    if len(organism_names) > 1:
        org_key = ' & '.join(organism_names)
        mixed_freqs = create_mixed_table([frequency_tables[org] for org in organism_names])
        frequency_tables[org_key] = mixed_freqs
        target_organism = org_key
    else:
        target_organism = organism_names[0]

    # Create AA to codon frequency mapping
    aa_codon_freq = create_aa_codon_frequency_table(frequency_tables[target_organism])

    # Encode sequences
    print(f"Encoding {len(sequences_df)} sequences using thresholded weighted method...")
    results = []
    for idx, row in sequences_df.iterrows():
        seq_id = row['id']
        protein_seq = row['sequence']

        dna_seq = encode_sequence_thresholded_weighted(protein_seq, aa_codon_freq)

        results.append({
            'id': seq_id,
            'protein_sequence': protein_seq,
            'dna_sequence': dna_seq,
            'organism': target_organism,
            'method': 'thresholded_weighted'
        })

    # Save CSV output
    results_df = pd.DataFrame(results)
    results_df.to_csv(dna_output, index=False)
    print(f"Saved DNA sequences to: {dna_output}")

    # Generate Excel output with color-coded codons
    print("Generating Excel file with color-coded codons...")
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DNA Sequences"

        # Headers
        headers = ["ID", "Protein", "Organism", "Method", "DNA Sequence (concatenated)"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        # Write sequences with color-coded codons
        row_idx = 2
        for idx, row in results_df.iterrows():
            ws.cell(row=row_idx, column=1, value=row['id'])
            ws.cell(row=row_idx, column=2, value=row['protein_sequence'])
            ws.cell(row=row_idx, column=3, value=row['organism'])
            ws.cell(row=row_idx, column=4, value=row['method'])

            # Split DNA into codons
            dna_seq = row['dna_sequence']
            codons = [dna_seq[i:i + 3] for i in range(0, len(dna_seq), 3)]

            # Write individual codons with color coding
            for col_offset, codon in enumerate(codons):
                cell = ws.cell(row=row_idx, column=6 + col_offset, value=codon)
                codon_freq = frequency_tables[target_organism].get(codon, 0)

                if codon_freq < 5:
                    cell.font = Font(color="FF0000")  # Red
                elif 5 <= codon_freq < 10:
                    cell.font = Font(color="FF9900")  # Orange
                # Default (black) for frequencies >= 10

            # Add concatenation formula for DNA sequence
            if len(codons) > 0:
                concat_refs = [f"{get_column_letter(6 + i)}{row_idx}" for i in range(len(codons))]
                concat_formula = f"=CONCATENATE({','.join(concat_refs)})"
                ws.cell(row=row_idx, column=5, value=concat_formula)

            row_idx += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 30
        for col in range(6, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = 5

        wb.save(excel_output)
        print(f"Saved Excel file to: {excel_output}")
    except ImportError:
        print("Warning: openpyxl not available, skipping Excel generation")
    except Exception as e:
        print(f"Warning: Failed to generate Excel file: {e}")

    # Write info file
    info_text = f"""DNA Encoding Information
========================

Method: Thresholded Weighted
- Codons with genome frequency >= 10‰ are sampled with relative weights
- If no codons >= 10‰, codons >= 5‰ are sampled
- If no codons >= 5‰, the most frequent codon is used

Organism(s): {target_organism}

Color coding in Excel:
- Red: codon frequency < 5‰
- Orange: codon frequency 5-10‰
- Black: codon frequency >= 10‰

Total sequences encoded: {len(results_df)}

Citation: Please cite CoCoPUTs (HIVE) when using these sequences.
Source: https://dnahive.fda.gov/dna.cgi?cmd=codon_usage
"""
    with open(info_output, 'w') as f:
        f.write(info_text)
    print(f"Saved encoding info to: {info_output}")

    print("DNA encoding completed successfully!")


if __name__ == '__main__':
    main()
